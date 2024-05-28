from flask import Flask, render_template, request, abort, jsonify, url_for
from jinja2 import TemplateNotFound
import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

app = Flask(__name__)
STOCKS_DIRECTORY = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'stocks', 'stockexcel')
SCREEN_XLSX_DIRECTORY = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'stocks', 'screenxlsx')

# Get the path of the directory where the script is located
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Load stock data from Excel files
def load_stock_data():
    stock_data = {}
    for filename in os.listdir(STOCKS_DIRECTORY):
        if filename.endswith('.xlsx'):
            stock_name = os.path.splitext(filename)[0]
            excel_file_path = os.path.join(STOCKS_DIRECTORY, filename)
            stock_data[stock_name] = excel_file_path  # Store the file path directly
    return stock_data

# Load stock data when the app starts
STOCK_DATA = load_stock_data()
# Function to retrieve Gainers, Losers, and Most Active stocks from Excel data
def get_stock_data(file_path):
    # Read the Excel file
    df = pd.read_excel(file_path)

    # Sort the data based on your criteria, such as price change or volume
    # Example: Sorting by percentage change in descending order
    sorted_data = df.sort_values(by='Column1.pChange', ascending=False)

    # Extract top gainers, losers, and most active stocks
    top_gainers = sorted_data.head(5)  # Assuming you want top 10 gainers
    top_losers = sorted_data.tail(5)    # Assuming you want top 10 losers
    most_active = sorted_data.sort_values(by='Column1.totalTradedVolume', ascending=False).head(10)  # Assuming you want top 10 most active

    # Extract stock symbols
    gainers = top_gainers['Column1.symbol'].tolist()
    losers = top_losers['Column1.symbol'].tolist()
    most_active_stocks = most_active['Column1.symbol'].tolist()

    return gainers, losers, most_active_stocks
# Example usage
gainers, losers, most_active = get_stock_data('stocks/stockexcel/nifty50.xlsx')
print("Gainers:", gainers)
print("Losers:", losers)
print("Most Active:", most_active)


# Main route
@app.route('/')
def index():
    stock_name = request.args.get('stock', None)
    if stock_name and stock_name in STOCK_DATA:
        excel_file_path = STOCK_DATA[stock_name]
        try:
            wb = load_workbook(excel_file_path, read_only=True, data_only=True)
            stock_data_html = {}
            ws = wb["Data Sheet"]  # Reading from "Data Sheet" sheet
            data = ws.values
            df = pd.DataFrame(data)
            stock_data_html["Data Sheet"] = df.to_html(index=False)
            
            # Define the data dictionary for Profit & Loss Statement
            data_profit_loss = {
                'Date': [cell.value.strftime('%b-%Y') if isinstance(cell.value, datetime) else cell.value for cell in ws['B16':'K16'][0]],
                'Sales': [ws.cell(row=17, column=i).value for i in range(2, 12)],
                'Expenses': [
                    round(
                        (ws.cell(row=18, column=i).value or 0) +
                        (ws.cell(row=20, column=i).value or 0) +
                        (ws.cell(row=21, column=i).value or 0) +
                        (ws.cell(row=22, column=i).value or 0) +
                        (ws.cell(row=23, column=i).value or 0) +
                        (ws.cell(row=24, column=i).value or 0) -
                        (ws.cell(row=19, column=i).value or 0),
                        2
                    )
                    for i in range(2, 12)
                ],
                'Operating Profit': [
                    round(
                        (ws.cell(row=17, column=i).value or 0) -
                        ((ws.cell(row=18, column=i).value or 0) +
                         (ws.cell(row=20, column=i).value or 0) +
                         (ws.cell(row=21, column=i).value or 0) +
                         (ws.cell(row=22, column=i).value or 0) +
                         (ws.cell(row=23, column=i).value or 0) +
                         (ws.cell(row=24, column=i).value or 0)) +
                        (ws.cell(row=19, column=i).value or 0),
                        2
                    )
                    for i in range(2, 12)
                ],
                'Other Income': [ws.cell(row=25, column=i).value for i in range(2, 12)],
                'Depreciation': [ws.cell(row=26, column=i).value for i in range(2, 12)],
                'Interest': [ws.cell(row=27, column=i).value for i in range(2, 12)],
                'Profit before tax': [ws.cell(row=28, column=i).value for i in range(2, 12)],
                'Tax': [ws.cell(row=29, column=i).value for i in range(2, 12)],
                'Net profit': [ws.cell(row=30, column=i).value for i in range(2, 12)],
                'EPS': [
                    round(
                        (ws.cell(row=30, column=i).value or 0) / (ws.cell(row=93, column=i).value or 1),
                        2
                    )
                    for i in range(2, 12)
                ],
                'Price to earning': [
                    round(
                        (ws.cell(row=90, column=i).value or 0) /
                        (
                            (ws.cell(row=30, column=i).value or 1) / (ws.cell(row=93, column=i).value or 1)
                        ),
                        2
                    )
                    if (ws.cell(row=90, column=i).value or 0) > 0
                    else ''
                    for i in range(2, 12)
                ],
                'Price': [ws.cell(row=90, column=i).value for i in range(2, 12)],
                'Dividend Payout': [
                    round(
                        (ws.cell(row=31, column=i).value or 0) / (ws.cell(row=30, column=i).value or 1) * 100,
                        2
                    )
                    for i in range(2, 12)
                ],
                'OPM': [
                    round(
                        ((ws.cell(row=17, column=i).value or 0) -
                         (ws.cell(row=18, column=i).value or 0) -
                         sum(ws.cell(row=row, column=i).value or 0 for row in range(20, 25)) +
                         (ws.cell(row=19, column=i).value or 0)) /
                        (ws.cell(row=17, column=i).value or 1) * 100,
                        2
                    )
                    for i in range(2, 12)
                ]
            }
            # Remove time from dates in data_profit_loss
            data_profit_loss['Date'] = [date.strftime('%b-%Y') if isinstance(date, datetime) else date for date in data_profit_loss['Date']]

            # Define data_cash_flow here
            data_cash_flow = {
                'Date': [cell.value.strftime('%b-%Y') if isinstance(cell.value, datetime) else cell.value for cell in ws['B81':'K81'][0]],
                'Cash from Operating Activity': [ws.cell(row=82, column=i).value for i in range(2, 12)],
                'Cash from Investing Activity': [ws.cell(row=83, column=i).value for i in range(2, 12)],
                'Cash from Financing Activity': [ws.cell(row=84, column=i).value for i in range(2, 12)],
                'Net Cash Flow': [ws.cell(row=85, column=i).value for i in range(2, 12)]
                # Add other data keys as needed
            }
            data_cash_flow['Date'] = [date.strftime('%b-%Y') if isinstance(date, datetime) else date for date in data_cash_flow['Date']]

            # Define the data dictionary for Quarters
            data_quarters = {
                'Date': [cell.value.strftime('%b-%Y') if isinstance(cell.value, datetime) else cell.value for cell in ws['B41':'K41'][0]],
                'Sales': [ws.cell(row=42, column=i).value for i in range(2, 12)],
                'Expenses': [ws.cell(row=43, column=i).value for i in range(2, 12)],
                'Operating Profit': [ws.cell(row=50, column=i).value for i in range(2, 12)],
                'Other Income': [ws.cell(row=44, column=i).value for i in range(2, 12)],
                'Depreciation': [ws.cell(row=45, column=i).value for i in range(2, 12)],
                'Interest': [ws.cell(row=46, column=i).value for i in range(2, 12)],
                'Profit Before Tax': [ws.cell(row=47, column=i).value for i in range(2, 12)],
                'Tax': [ws.cell(row=48, column=i).value for i in range(2, 12)],
                'Net Profit': [ws.cell(row=49, column=i).value for i in range(2, 12)]
            }
            data_quarters['Date'] = [date.strftime('%b-%Y') if isinstance(date, datetime) else date for date in data_quarters['Date']]

                # Define the data dictionary for Balance Sheet
            data_balance_sheet = {
                'Date':[cell.value.strftime('%b-%Y') if isinstance(cell.value, datetime) else cell.value for cell in ws['B56':'K56'][0]],
                'Equity Share Capital': [ws.cell(row=57, column=i).value for i in range(2, 12)],
                'Reserves': [ws.cell(row=58, column=i).value for i in range(2, 12)],
                'Borrowings': [ws.cell(row=59, column=i).value for i in range(2, 12)],
                'Other Liabilities': [ws.cell(row=60, column=i).value for i in range(2, 12)],
                'Total': [ws.cell(row=61, column=i).value for i in range(2, 12)],
                'Net Block': [ws.cell(row=62, column=i).value for i in range(2, 12)],
                'Capital Work in Progress': [ws.cell(row=63, column=i).value for i in range(2, 12)],
                'Investments': [ws.cell(row=64, column=i).value for i in range(2, 12)],
                'Other Assets': [ws.cell(row=65, column=i).value for i in range(2, 12)],
                'Total Assets': [ws.cell(row=66, column=i).value for i in range(2, 12)],
                'Working Capital': [
                    round(
                        (ws.cell(row=65, column=i).value or 0) - (ws.cell(row=60, column=i).value or 0),
                        2
                    )
                    for i in range(2, 12)
                ],
                'Debtors': [ws.cell(row=67, column=i).value for i in range(2, 12)],
                'Inventory': [ws.cell(row=68, column=i).value for i in range(2, 12)],
                'Debtor Days': [
                    round((ws.cell(row=67, column=i).value or 0) / ((ws.cell(row=17, column=i).value or 1) / 365), 2)
                       if (ws.cell(row=17, column=i).value or 0) > 0 else 0
                        for i in range(2, 12)
                ],
                'Inventory Turnover': [
                    round(
                        (ws.cell(row=17, column=i).value or 0) / (ws.cell(row=68, column=i).value or 1),
                        2
                    )
                    if (ws.cell(row=68, column=i).value or 0) > 0 else 0
                    for i in range(2, 12)
                ],
                'Return on Equity': [
                    f"{round(((ws.cell(row=30, column=i).value or 0) / 
                             (sum(ws.cell(row=row, column=i).value or 0 for row in range(57, 59)))) * 100, 2)}%"
                    if sum(ws.cell(row=row, column=i).value or 0 for row in range(57, 59)) > 0 else "-"
                    for i in range(2, 12)
                ],
                'Return on Capital Employed': [
                    '-' if i == 2 else  # Print "-" in the first column
                    f"{round(((ws.cell(row=28, column=i).value or 0) + (ws.cell(row=27, column=i).value or 0)) * 2 /
                             (sum((ws.cell(row=row, column=i-1).value or 0) for row in range(57, 60)) +
                              sum((ws.cell(row=row, column=i).value or 0) for row in range(57, 60))) * 100, 2)}%"
                    if (sum((ws.cell(row=row, column=i-1).value or 0) for row in range(57, 60)) +
                        sum((ws.cell(row=row, column=i).value or 0) for row in range(57, 60))) > 0
                    else "-"
                    for i in range(2, 12)  # Starting from the second column (index 2)
                ]
            }
                        # Remove time from dates in data_profit_loss
            data_balance_sheet['Date'] = [date.strftime('%b-%Y') if isinstance(date, datetime) else date  for date in data_balance_sheet['Date']]
            # Update other_data dictionary
            other_data = {
                'Facevalue': ws.cell(row=7, column=2).value,
                'Currentprice': ws.cell(row=8, column=2).value,
                'Marketcap': ws.cell(row=9, column=2).value
            }
            return render_template('stock_analysis.html', stock_name=stock_name, stock_data_html=stock_data_html,data_profit_loss=data_profit_loss,data_cash_flow=data_cash_flow,data_quarters=data_quarters,data_balance_sheet=data_balance_sheet,other_data=other_data)
        except Exception as e:
            return render_template('error.html', message=f'Error loading data for {stock_name}: {str(e)}')
    else:
        return render_template('index.html', gainers=gainers, losers=losers, most_active=most_active)



# Autocomplete route
@app.route('/autocomplete')
def autocomplete():
    query = request.args.get('query', '')
    suggestions = [{'name': stock_name, 'url': url_for('view_stock', stock_name=stock_name)} for stock_name in STOCK_DATA.keys() if query.lower() in stock_name.lower()]
    suggestions.sort(key=lambda x: x['name'])
    suggestions = suggestions[:5]  # Limit the number of suggestions
    return jsonify(suggestions)


# Route to display the stock analysis
@app.route('/stocks/<stock_name>')
def view_stock(stock_name):
    print("Viewing stock:", stock_name)  # Add this line for debugging
    if stock_name in STOCK_DATA:
        excel_file_path = STOCK_DATA[stock_name]
        print("Excel file path:", excel_file_path)  # Add this line for debugging
        try:
            wb = load_workbook(excel_file_path)
            stock_data_html = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                data = ws.values
                df = pd.DataFrame(data)
                stock_data_html[sheet_name] = df.to_html(index=False, header=False)
            
            # Additional data for stock analysis
            data_quarters = {}  # You need to define this data
            data_cash_flow = {}  # You need to define this data
            data_balance_sheet = {}  # You need to define this data

            return render_template('stock_analysis.html', stock_name=stock_name, stock_data_html=stock_data_html, data_quarters=data_quarters, data_cash_flow=data_cash_flow, data_balance_sheet=data_balance_sheet)
        except Exception as e:
            return render_template('error.html', message=f'Error loading data for {stock_name}: {str(e)}')
    else:
        abort(404)



# Route to serve navbar.html
@app.route('/static/templates/navbar.html')
def serve_navbar():
    return render_template('navbar.html')

# Route to serve sector
@app.route('/sector/<sector_name>/index.html')
def sector(sector_name):
    return f"You selected the sector: {sector_name}"

# Route to serve all screens
@app.route('/screens/all-screens.html')
def explore_screener():
    return render_template('/screens/all-screens.html')

# Route to serve individual screens HTML
@app.route('/screens/indscreenshtml/<filename>')
def serve_ind_screens_html(filename):
    try:
        # Strip the .html extension from the filename
        base_filename = os.path.splitext(filename)[0]

        # Construct the Excel file path
        excel_file_path = os.path.join(SCREEN_XLSX_DIRECTORY, f'{base_filename}.xlsx')

        # Check if the Excel file exists
        if not os.path.exists(excel_file_path):
            raise FileNotFoundError(f'Excel file not found: {excel_file_path}')

        # Load Excel data
        df = pd.read_excel(excel_file_path)
        data_html = df.to_html(index=False)

        # Check if the template file exists
        template_path = f'screens/indscreenshtml/{filename}'
        if not os.path.exists(os.path.join(app.template_folder, template_path)):
            raise FileNotFoundError(f'Template file not found: {template_path}')

        return render_template(template_path, data_html=data_html)
    except FileNotFoundError as e:
        return render_template('error.html', message=str(e)), 404
    except Exception as e:
        return render_template('error.html', message=f'Error fetching data for {filename}: {str(e)}'), 500
    
# Route to serve financial calculator HTML files
@app.route('/finance/<path:filename>')
def serve_finance_html(filename):
    try:
        return render_template(f'finance/{filename}')
    except TemplateNotFound:
        abort(404)

# Route for contact page
@app.route('/contact.html')
def contact():
    return render_template('contact.html')

# Route for login page
@app.route('/login.html')
def login():
    return render_template('login.html')

# Route for registration page
@app.route('/register.html')  
def register():
    return render_template('register.html') 

# Error handling route
@app.errorhandler(404)
def page_not_found(error):
    return render_template('error.html', message='Page not found'), 404

if __name__ == '__main__':
    app.run(debug=True)
